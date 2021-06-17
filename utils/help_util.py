import json
from openpyxl import load_workbook
from utilities.models import Protocol


class Helper:
	def __init__(self, filename='data/Protocols.xlsx',model_name=None):
		self.filename = filename
		self.model_name = model_name
		self.wb = load_workbook(filename)
		self.model_helpers = {}
		self._set_values()

	def _set_values(self):
		for sheet_name in self.wb.sheetnames:
			if self.model_name != None and sheet_name.lower() != self.model_name.lower():continue
			self.model_helpers[sheet_name] = ModelHelper(self.wb,sheet_name)

	def get_dict(self,name=None):
		if self.model_name != None:name = self.model_name
		for key in self.model_helpers.keys():
			if name.lower() == key.lower(): 
				model_helper = self.model_helpers[key]
				break
		return model_helper._make_dict()

	def get_json(self,name=None):
		d = self.get_dict(name)
		return json.dump(d)
		

class ModelHelper:
	def __init__(self,wb,name):
		self.protocols = Protocol.objects.filter(model_name=name.lower())
		self.wb = wb
		self.name = name
		self.sheet = wb[name]
		self.field_helpers = {}
		self._set_values()

	def _set_values(self):
		self.values = list(self.sheet.values)
		self.errors = []
		for line in self.values:
			if len(line) < 2 or line[1] == 'Field':continue
			fh = FieldHelper(line)
			if fh.ok: self.field_helpers[fh.field_name] = fh
			else: self.errors.append(line)
				
	def _make_dict(self):
		d = {}
		if self.protocols:
			print('creating help text with protocols')
			for p in self.protocols:
				d[p.field_name] =  p.explanation 
				d[p.field_name.lower().strip()] = '<p>' + p.explanation + '</p>'
		else:
			for fh in self.field_helpers.values():
				d[fh.field_name] =fh.help_text
				d[fh.field_name.lower().strip()] =fh.help_text_html
		return d
			
		

class FieldHelper:
	def __init__(self,line):
		self.line = line
		self.ok = True
		try:
			self.model_name, self.field_name, self.help_text = line[:3]
			if self.field_name == None: self.ok = False
		except:self.ok = False
		if self.ok:
			self.id = self.field_name.lower().replace(' ','_')
			self.help_text_html = ''
			for line in self.help_text.split('\n'):
				self.help_text_html += '<p>'+line+'</p> '
			

	def __repr__(self):
		return self.model_name + ' | ' + self.id

	def __str__(self):
		m = self.__repr__()
		m +='\n' + self.help_text
		return m
	
names = 'Image,Infographic,Film,Music,PictureStory,Text,Videogame,Recordedspeech'
names += ',Memorialsite,Artefact'
other_names = 'Famine,Location,Person,Keyword'.split(',')
app_names  = 'misc,locations,persons,misc'.split(',')

names = ','.join(['sources|'+n for n in names.split(',')]) +','
names += ','.join([app_names[i] +'|' +n for i,n in enumerate(other_names)])

def load_protocol_in_database(save= False):
	h = Helper()
	protocols = []
	for name in names.split(','):
		print(name)
		app_name, model_name = name.split('|')
		d = h.get_dict(model_name)
		for field_name in d.keys():
			if '<p>' in d[field_name]: continue
			explanation = d[field_name]
			p = _make_protocol(app_name,model_name,field_name,explanation,save)
			protocols.append(p)
	print('done')
	return protocols

def _make_protocol(app_name,model_name,field_name,explanation,save=False):
	app_name, model_name = app_name.lower(), model_name.lower()
	try:p = Protocol.objects.get(app_name=app_name,model_name=model_name,field_name=field_name)
	except: 
		print('making new protocol:',model_name,field_name,'\n',explanation,'\n---\n')
		p = Protocol(app_name=app_name,model_name=model_name,
			field_name=field_name, explanation= explanation)
	else:
		print('setting explanation (protocol exists):',p.explanation,'\nwith\n',explanation,'\n---\n')
		p.explanation = explanation
	if save:p.save()
	return p
	
		
	
		
