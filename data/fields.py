from sources.models import Text, Image, Film, Memorialsite, Music
from sources.models import Artefact, PictureStory, Recordedspeech 
from sources.models import Infographic, Videogame
from persons.models import Person

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Fields:
    def __init__(self):
        self.models = [Text, Image, Film, Memorialsite, Music, 
            Person, Artefact, PictureStory, Recordedspeech, Infographic, 
            Videogame]
        self._set_field_names()

    def _set_field_names(self):
        self.all_categories = all_categories
        self.text_fields = text
        self.image_fields = image
        self.film_fields = film
        self.memorialsite_fields = memorialsite
        self.music_fields = music
        self.person_fields = person
        self.artefact_fields = artefact
        self.picturestory_fields = picturestory
        self.recordedspeech_fields = recordedspeech
        self.infographic_fields = infographic
        self.videogame_fields = videogame

    def check_instances(self, model):
        model_name = model._meta.model_name
        instances = model.objects.all()
        if model_name == 'person': 
            instances = [x for x in instances if x.flag]
        instances_with_empty_fields = []
        for instance in instances:
            empty_fields = self.check_instance(instance)
            if len(empty_fields) > 0: 
                instances_with_empty_fields.append([instance, empty_fields])
        print(model._meta.model_name.ljust(16) ,str(len(instances)).ljust(5), 
            len(instances_with_empty_fields))
        return instances_with_empty_fields

    def check_instance(self, instance):
        model = instance._meta.model
        field_names = self._get_field_names(model)
        empty_fields = []
        for field_name in field_names:
            ok, field_type = self._check_field(model, field_name)
            if field_type == 'field' or field_type == 'other':
                field = getattr(instance, field_name)
                if not field: empty_fields.append(field_name)
            elif field_type == 'm2m':
                queryset= getattr(instance, field_name).all()
                if not queryset.count(): empty_fields.append(field_name)
            else: 
                raise ValueError('field_type not found',field_type,field_name)
        return empty_fields

    def _get_field_names(self, model):
        model_name = model._meta.model_name
        return getattr(self, model_name + '_fields')

    def _check_field(self, model, field_name):
        ok = False
        for field in model._meta.fields:
            if field.name == field_name: 
                ok = True
                field_type = 'field'
        
        if not ok and hasattr(model, field_name): 
            ok = True
            field = getattr(model, field_name)
            if hasattr(field, 'all'):
                field_type = 'm2m'
            else: 
                field_type = 'other'
        return ok, field_type

    def _check_fields_exists(self, model, field_names):
        error = []
        for field_name in field_names:
            ok, field_type = self._check_field(model, field_name)
            if not ok:
                error.append(field_name)
        if len(error) > 0: 
            print(model._meta.model_name, 'error:', error)
        else:
            print(model._meta.model_name, 'all ok')

    def _check_fields_for_all_models(self):
        for model in self.models:
            field_names = self._get_field_names(model)
            self._check_fields_exists(model, field_names)

    def _add_identifier_column(self, empties, sheet):
        add_cell_value(sheet, 1, 1, 'identifier')
        for row_index, empty in enumerate(empties):
            instance = empty[0]
            add_cell_value(sheet, row_index+2, 1, instance.identifier)
        sheet.column_dimensions['A'].width = 18

    def _add_field_column(self, empties, sheet):
        add_cell_value(sheet, 1, 2, 'empty_fields')
        for row_index, empty in enumerate(empties):
            field_names= ', '.join(empty[1])
            add_cell_value(sheet, row_index+2, 2, field_names)
        sheet.column_dimensions['b'].width = 50

    def _add_url_column(self, empties, sheet):
        add_cell_value(sheet, 1, 3, 'url')
        for row_index, empty in enumerate(empties):
            instance = empty[0]
            add_cell_value(sheet, row_index+2, 3, 'link',
                instance.edit_url_complete)

    def _model_to_excel(self, model):
        sheet = add_sheet(self.wb, model._meta.model_name)
        empties = self.check_instances(model)
        self._add_identifier_column(empties, sheet)
        self._add_field_column(empties, sheet)
        self._add_url_column(empties, sheet)

    def to_excel(self, filename= 'data/empty_fields.xlsx'):
        self.wb = make_workbook()
        for model in self.models:
            self._model_to_excel(model)
        remove_sheet(self.wb, 'Sheet')
        self.wb.save(filename)
        return self.wb

def make_workbook():
    return Workbook()

def add_sheet(wb, sheet_name):
    return wb.create_sheet(sheet_name)

def remove_sheet(wb, sheet_name):
    wb.remove(wb[sheet_name])

def add_cell_value(sheet, row, column, value, hyperlink = ''):
    cell = sheet.cell(row=row,column=column,value= value)
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.style = 'Hyperlink'
    return cell



all_categories = 'description,famines,keywords,links,rated,license_image'
all_categories += ',license_thumbnail'
all_categories = all_categories.split(',')
text = 'authors,languages,setting,date_released,text_type'
text = text.split(',')
text.extend(all_categories)
image = 'creators,setting,date_created,image_type'
image = image.split(',')
image.extend(all_categories)
film = 'creators,languages_original,date_released,film_type'
film = film.split(',')
film.extend(all_categories)
memorialsite = 'creators,locations,date_released'
memorialsite = memorialsite.split(',')
memorialsite.extend(all_categories)
music = 'composers,languages,date_released,music_type'
music = music.split(',')
music.extend(all_categories)
person = 'date_of_birth,date_of_death,biography_link,occupation'
person = person.split(',')
person.extend(['description', 'famines', 'keywords'])
artefact = 'locations,date_created,artefact_type'
artefact = artefact.split(',')
artefact.extend(all_categories)
picturestory = 'authors,artists,languages,date_released,picture_story_type'
picturestory = picturestory.split(',')
picturestory.extend(all_categories)
recordedspeech = 'speakers,languages,date_released,recordedspeech_type'
recordedspeech = recordedspeech.split(',')
recordedspeech.extend(all_categories)
infographic = 'creators,languages,date_created,infographic_type'
infographic = infographic.split(',')
infographic.extend(all_categories)
videogame = 'production_studio,date_released,game_type' 
videogame = videogame.split(',')
videogame.extend(all_categories)
